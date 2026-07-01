import os
import re
import unicodedata
from calendar import monthrange
from openpyxl import load_workbook

TARGET_BLOCKS = [
    "玉野",
    "現金機＆CLAP"
]


def clean(text):
    if text is None:
        return ""
    return (
        unicodedata.normalize("NFKC", str(text))
        .replace("　", "")
        .replace(" ", "")
        .strip()
        .replace("&", "＆")
    )


def build_merged_map(ws):
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        value = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = value
    return merged_map


def merged_value(cell, merged_map):
    if cell.value is not None:
        return cell.value
    return merged_map.get((cell.row, cell.column))


def find_months(ws, merged_map):
    months = {}

    for row in ws.iter_rows():
        for cell in row:
            value = clean(merged_value(cell, merged_map))

            m = re.match(r"^([1-9]|1[0-2])月$", value)
            if m:
                month = int(m.group(1))
                if month not in months:
                    months[month] = cell

    return months


def get_day1_column(ws, month_cell):
    for merged in ws.merged_cells.ranges:
        if month_cell.coordinate in merged:
            return merged.max_col + 1
    return month_cell.column + 1


def resolve_year(filename, month):
    m = re.search(r"R([0-9０-９]+)", filename)

    if m:
        fiscal = 2018 + int(
            m.group(1).translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        )
    else:
        m = re.search(r"(20\d{2})", filename)
        fiscal = int(m.group(1))

    return fiscal if month >= 4 else fiscal + 1


def find_block_column(ws, merged_map):
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, 10):
            value = clean(merged_value(ws.cell(row, col), merged_map))
            if value in TARGET_BLOCKS:
                return col
    return 2


def get_place_name_from_excel(excel_path, target_date):
    """
    戻り値
        玉野市営玉野競輪
        高松市営玉野競輪
        小松島市営玉野競輪
    """

    wb = load_workbook(excel_path, data_only=True)
    filename = os.path.basename(excel_path)

    try:
        for ws in wb.worksheets:

            if ws.sheet_state != "visible":
                continue

            merged_map = build_merged_map(ws)
            month_map = find_months(ws, merged_map)

            if target_date.month not in month_map:
                continue

            month_cell = month_map[target_date.month]

            if resolve_year(filename, target_date.month) != target_date.year:
                continue

            block_col = find_block_column(ws, merged_map)
            day1_col = get_day1_column(ws, month_cell)
            target_col = day1_col + target_date.day - 1

            start_row = month_cell.row
            end_row = ws.max_row

            # 次の月の開始行まで
            sorted_months = sorted(
                month_map.items(),
                key=lambda x: (x[1].row, x[1].column)
            )

            for i, (_, cell) in enumerate(sorted_months):
                if cell == month_cell:
                    if i + 1 < len(sorted_months):
                        end_row = sorted_months[i + 1][1].row - 1
                    break

            current_block = None

            for row in range(start_row, end_row + 1):

                block = clean(
                    merged_value(ws.cell(row, block_col), merged_map)
                )

                if block:
                    if block in TARGET_BLOCKS:
                        current_block = block
                    else:
                        current_block = None

                if current_block != "玉野":
                    continue

                value = clean(
                    merged_value(ws.cell(row, target_col), merged_map)
                )

                if not value:
                    continue

                # 高松in玉野・小松島in玉野
                m = re.match(r"(.+?)in玉野$", value)
                if m:
                    return f"{m.group(1)}市営玉野競輪"

                # 通常開催
                if value == "玉野":
                    return "玉野市営玉野競輪"

            break

    finally:
        wb.close()

    return "玉野市営玉野競輪"
